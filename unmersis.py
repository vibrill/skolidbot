from openpyxl import Workbook
from openpyxl import load_workbook

def unmersis(path):
    wb = load_workbook(filename = path)

    # grab the active worksheet
    ws = wb.active
    #---------------------------------------
    for i in range(ord('A'), ord('X')+1):
        try:
            #print (chr(i))
            ws.unmerge_cells(chr(i)+'5:'+chr(i)+'6')
        except:
            print('err di '+ chr(i)+'5:'+chr(i)+'6')
    #---------------------------------------    
    try:
        ws.unmerge_cells('Y5:AD5')
    except:
        print('err di Y5:AD5')
    #---------------------------------------   
    try:
        ws.unmerge_cells('AE5:AJ5')
    except:
        print('err di AE5:AJ5')
    #---------------------------------------    
    try:
        ws.unmerge_cells('AK5:AP5')
    except:
        print('err di AK5:AP5')
    #---------------------------------------
    for i in range(ord('Q'), ord('Z')+1):
        try:
            ws.unmerge_cells('A'+chr(i)+'5:'+'A'+chr(i)+'6')
        except:
            print('err di '+'A'+chr(i)+'5:'+'A'+chr(i)+'6')
    #---------------------------------------
    for i in range(ord('A'), ord('N')+1):
        try:
            ws.unmerge_cells('B'+chr(i)+'5:'+'B'+chr(i)+'6')
        except:
            print('err di '+'B'+chr(i)+'5:'+'B'+chr(i)+'6')
    #---------------------------------------
    # Save the file
    wb.save(path)

def delearn(path):
    wb = load_workbook(filename = path)
    ws = wb.active
    ws.delete_cols(29)
    ws.delete_cols(34)
    ws.delete_cols(39)
    wb.save(path)

# Python types will automatically be converted
#import datetime
#ws['A2'] = datetime.datetime.now()
# Data can be assigned directly to cells
#ws['A1'] = 42
# Rows can also be appended
#ws.append([1, 2, 3])
#ws.delete_rows(5, amount=1)
#for i in range(ord('a'), ord('z')+1):
#    print (chr(i))
